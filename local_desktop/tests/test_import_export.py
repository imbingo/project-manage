import json
import sys
from pathlib import Path

import pytest

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from openpyxl import load_workbook

from src.import_export import dump_workspace_json, export_project_excel, export_tasks_csv, is_gantt_truncated, load_workspace_json, normalize_workspace
from src.models import DailyLog, ProgressEntry, Project, Task, Workspace


def test_import_workspace_json():
    workspace, diagnostics = normalize_workspace(
        {
            "selectedProjectId": "p1",
            "selectedDate": "2026-05-06",
            "projects": [
                {
                    "id": "p1",
                    "name": "项目",
                    "deadline": "2026-05-30",
                    "tasks": [{"id": "t1", "title": "任务", "startDate": "2026-05-06", "duration": 4}],
                }
            ],
        }
    )
    assert workspace.selectedProjectId == "p1"
    assert workspace.projects[0].tasks[0].duration == 4
    assert diagnostics[0].startswith("识别格式")


def test_import_legacy_payload_and_snake_case():
    workspace, _ = normalize_workspace(
        {
            "payload": {
                "projects": [
                    {
                        "id": "p1",
                        "top_risk": "风险",
                        "next_step": "计划",
                        "tasks": [
                            {
                                "id": "t1",
                                "parent_id": None,
                                "risk": "X",
                                "title": "任务",
                                "responsible": "张三",
                                "start_date": "2026-05-06",
                                "duration": "0",
                                "status": "Bad",
                                "progress_entries": [{"entry_date": "2026-05-06", "planned_progress": 120, "actual_progress": -1}],
                            }
                        ],
                        "daily_logs": [
                            {
                                "task_id": "t1",
                                "log_date": "2026-05-06",
                                "plan_text": "计划",
                                "actual_text": "实际",
                                "planned_progress": 50,
                                "actual_progress": 40,
                            }
                        ],
                    }
                ]
            }
        }
    )
    project = workspace.projects[0]
    task = project.tasks[0]
    assert project.topRisk == "风险"
    assert project.nextStep == "计划"
    assert task.risk == "M"
    assert task.status == "Open"
    assert task.duration == 1
    assert task.progressEntries[0].plannedProgress == 100
    assert task.progressEntries[0].actualProgress == 0
    assert project.dailyLogs[0].planText == "计划"


def test_import_local_storage_wrapped_json_string():
    payload = {"projects": [{"id": "p1", "name": "项目", "tasks": []}]}
    workspace, diagnostics = normalize_workspace({"project-desk-local-v5": json.dumps(payload, ensure_ascii=False)})
    assert workspace.projects[0].name == "项目"
    assert diagnostics[0] == "识别格式：localStorage project-desk-local-v5"


def test_export_excel(tmp_path):
    task = Task(
        id="t1",
        title="任务",
        responsible="李四",
        startDate="2026-05-06",
        duration=4,
        progressEntries=[ProgressEntry(entryDate="2026-05-06", plannedProgress=25, actualProgress=10)],
    )
    project = Project(
        id="p1",
        name="项目",
        deadline="2026-05-30",
        summary="一句话",
        topRisk="风险",
        nextStep="计划",
        tasks=[task],
        dailyLogs=[DailyLog(taskId="t1", date="2026-05-06", responsible="李四", planText="计划", actualText="实际")],
    )
    path = tmp_path / "project.xlsx"
    export_project_excel(project, path)
    assert path.exists()
    assert path.stat().st_size > 1000


def test_export_excel_marks_truncated_gantt(tmp_path):
    task = Task(id="t1", title="长周期任务", startDate="2026-01-01", duration=150)
    project = Project(id="p1", name="长项目", tasks=[task])
    path = tmp_path / "long.xlsx"
    assert is_gantt_truncated(project)
    export_project_excel(project, path)
    workbook = load_workbook(path)
    values = [cell.value for row in workbook.active.iter_rows() for cell in row if cell.value]
    assert any("超过 120 天" in str(value) for value in values)


def test_json_and_csv_export_round_trip(tmp_path):
    task = Task(id="t1", title="任务", responsible="张三", startDate="2026-05-06", duration=4)
    project = Project(id="p1", name="项目", deadline="2026-05-30", tasks=[task])
    workspace = Workspace(selectedProjectId="p1", selectedDate="2026-05-06", projects=[project])
    json_path = tmp_path / "workspace.json"
    csv_path = tmp_path / "tasks.csv"
    dump_workspace_json(workspace, json_path)
    loaded, diagnostics = load_workspace_json(json_path)
    assert loaded.projects[0].name == "项目"
    assert diagnostics[0].startswith("识别格式")
    export_tasks_csv(project, csv_path)
    content = csv_path.read_text(encoding="utf-8-sig")
    assert "任务" in content
    assert "张三" in content
