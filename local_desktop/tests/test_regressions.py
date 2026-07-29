import json
import os
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))
os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")

from openpyxl import load_workbook
from PySide6.QtCore import Qt
from PySide6.QtWidgets import QApplication, QCalendarWidget, QComboBox, QDateEdit, QPushButton, QStackedWidget, QTableWidget

import src.app as app_module
from src.app import AppDialog, ArchiveDialog, DailyDialog, InboxTaskDialog, MainWindow, ProjectDialog, TaskDialog, app_icon, ordered_tasks, sorted_projects
from src.import_export import dump_workspace_json, export_project_excel, export_tasks_csv, load_workspace_json, normalize_workspace
from src.metrics import normalize_date, task_end_date
from src.models import APP_VERSION, DailyLog, ProgressEntry, Project, Task, Workspace
from src.operations import delete_daily_log, save_daily_log, update_task
from src.storage import load_workspace, save_workspace, workspace_path


def app():
    return QApplication.instance() or QApplication([])


def make_cross_project_workspace():
    task_a = Task(id="t-a", title="Alpha task", responsible="Owner A", startDate="2026-06-01", duration=2)
    task_b = Task(id="t-b", title="Beta task", responsible="Owner B", startDate="2026-06-03", duration=3, risk="H")
    project_a = Project(
        id="p-a",
        name="Alpha project",
        deadline="2026-06-30",
        tasks=[task_a],
        dailyLogs=[DailyLog(id="l-a", taskId="t-a", date="2026-06-02", responsible="Owner A", planText="Plan A", actualText="Actual A")],
    )
    project_b = Project(
        id="p-b",
        name="Beta project",
        deadline="2026-07-15",
        tasks=[task_b],
        dailyLogs=[
            DailyLog(
                id="l-b",
                taskId="t-b",
                date="2026-06-04",
                responsible="Owner B",
                planText="Plan B",
                actualText="Actual B",
                result="延期",
                delayReason="Blocked",
            )
        ],
    )
    return Workspace(selectedProjectId="p-b", selectedDate="2026-06-04", projects=[project_a, project_b])


def window_for_workspace(monkeypatch, workspace):
    app()
    monkeypatch.setattr(app_module, "load_workspace", lambda: workspace)
    monkeypatch.setattr(app_module, "save_workspace", lambda _workspace: None)
    return MainWindow()


def table_with_headers(page, headers):
    for table in page.findChildren(QTableWidget):
        current = [table.horizontalHeaderItem(index).text() for index in range(min(table.columnCount(), len(headers)))]
        if current == headers:
            return table
    raise AssertionError(f"table with headers {headers} not found")


def test_daily_overview_date_range_includes_start_date():
    task = Task(title="range", startDate="2026-06-02", duration=8, status="Open")
    selected = normalize_date("6.2", fallback_year=2026)
    start = normalize_date(task.startDate, fallback_year=2026)
    end = task_end_date(task)

    assert selected == "2026-06-02"
    assert start <= selected <= end


def test_delete_completion_daily_log_recalculates_task_status():
    workspace = Workspace(projects=[])
    project = Project(name="p")
    task = Task(title="t", startDate="2026-06-01", duration=3)
    project.tasks = [task]
    workspace.projects = [project]

    save_daily_log(
        workspace,
        project,
        None,
        {
            "taskId": task.id,
            "date": "2026-06-01",
            "responsible": "a",
            "planText": "p",
            "actualText": "a",
            "plannedProgress": 50,
            "actualProgress": 50,
            "result": "部分完成",
            "delayReason": "",
        },
    )
    done = save_daily_log(
        workspace,
        project,
        None,
        {
            "taskId": task.id,
            "date": "2026-06-02",
            "responsible": "a",
            "planText": "p",
            "actualText": "a",
            "plannedProgress": 100,
            "actualProgress": 100,
            "result": "完成",
            "delayReason": "",
        },
    )

    assert task.status == "Closed"
    delete_daily_log(project, done.id)

    assert task.status == "Ongoing"
    assert task.completedDate == ""


def test_update_task_does_not_write_progress_when_unchanged():
    task = Task(
        title="Task",
        startDate="2026-06-01",
        progressEntries=[
            ProgressEntry(entryDate="2026-06-01", plannedProgress=10, actualProgress=10),
            ProgressEntry(entryDate="2026-06-09", plannedProgress=80, actualProgress=80),
        ],
    )

    update_task(task, "2026-06-02", {"title": "Renamed", "plannedProgress": 80, "actualProgress": 80})

    assert task.title == "Renamed"
    assert [(entry.entryDate, entry.plannedProgress, entry.actualProgress) for entry in task.progressEntries] == [
        ("2026-06-01", 10, 10),
        ("2026-06-09", 80, 80),
    ]


def test_update_task_writes_progress_only_when_progress_changed():
    task = Task(
        title="Task",
        startDate="2026-06-01",
        progressEntries=[ProgressEntry(entryDate="2026-06-01", plannedProgress=10, actualProgress=10)],
    )

    update_task(task, "2026-06-02", {"plannedProgress": 20, "actualProgress": 15})

    assert [(entry.entryDate, entry.plannedProgress, entry.actualProgress) for entry in task.progressEntries] == [
        ("2026-06-01", 10, 10),
        ("2026-06-02", 20, 15),
    ]


def test_import_duration_accepts_float_like_text():
    workspace, _ = normalize_workspace(
        {
            "projects": [
                {
                    "name": "p",
                    "tasks": [
                        {
                            "title": "task",
                            "startDate": "2026-06-01",
                            "duration": "4.0",
                        }
                    ],
                }
            ]
        }
    )

    assert workspace.projects[0].tasks[0].duration == 4


def test_storage_writes_current_app_version(tmp_path, monkeypatch):
    monkeypatch.setenv("APPDATA", str(tmp_path))
    workspace = Workspace(projects=[Project(id="p1", name="Project")], selectedProjectId="p1")

    save_workspace(workspace)

    payload = json.loads(workspace_path().read_text(encoding="utf-8"))
    assert payload["version"] == APP_VERSION
    assert payload["workspace"]["version"] == APP_VERSION
    assert load_workspace().version == APP_VERSION


def test_app_version_is_v3_5():
    assert APP_VERSION == "Project_Manage_LocalV3.5"


def test_application_icon_is_available():
    app()

    icon = app_icon()

    assert not icon.isNull()


def test_completed_tasks_are_sorted_after_open_tasks():
    done = Task(id="done", title="Done", startDate="2026-06-01", status="Closed")
    open_task = Task(id="open", title="Open", startDate="2026-06-09", status="Open")

    ordered = ordered_tasks([done, open_task])

    assert [task.id for task, _depth in ordered] == ["open", "done"]


def test_completed_projects_are_sorted_after_active_projects(monkeypatch):
    active = Project(id="active", name="Active", deadline="2026-07-30", tasks=[Task(status="Open")])
    done = Project(id="done", name="Done", deadline="2026-06-01", tasks=[Task(status="Closed")])
    workspace = Workspace(selectedProjectId="active", projects=[done, active])

    assert [project.id for project in sorted_projects(workspace.projects)] == ["active", "done"]

    window = window_for_workspace(monkeypatch, workspace)
    combo_values = [window.project_select.itemData(index) for index in range(window.project_select.count())]

    assert combo_values == ["active", "done"]


def test_json_csv_excel_export_round_trip(tmp_path):
    task = Task(
        id="t1",
        title="Task",
        responsible="Owner",
        startDate="2026-06-01",
        duration=4,
        progressEntries=[ProgressEntry(entryDate="2026-06-01", plannedProgress=20, actualProgress=10)],
    )
    project = Project(
        id="p1",
        name="Project",
        deadline="2026-06-30",
        summary="summary",
        topRisk="risk",
        nextStep="next",
        tasks=[task],
        dailyLogs=[DailyLog(taskId="t1", date="2026-06-01", responsible="Owner", planText="Plan", actualText="Actual")],
    )
    workspace = Workspace(selectedProjectId="p1", selectedDate="2026-06-01", projects=[project])

    json_path = tmp_path / "workspace.json"
    csv_path = tmp_path / "tasks.csv"
    xlsx_path = tmp_path / "project.xlsx"
    dump_workspace_json(workspace, json_path)
    loaded, diagnostics = load_workspace_json(json_path)
    export_tasks_csv(project, csv_path)
    export_project_excel(project, xlsx_path)

    assert loaded.projects[0].name == "Project"
    assert diagnostics[0].startswith("识别格式")
    assert "Task" in csv_path.read_text(encoding="utf-8-sig")
    workbook = load_workbook(xlsx_path)
    values = [cell.value for row in workbook.active.iter_rows() for cell in row if cell.value]
    assert "Project" in values
    assert "Task" in values


def test_core_date_fields_use_calendar_widgets():
    app()
    project = Project(
        name="Project",
        deadline="2026-06-30",
        tasks=[Task(id="t1", title="Task", startDate="2026-06-01", completedDate="2026-06-03")],
    )

    project_dialog = ProjectDialog(project)
    task_dialog = TaskDialog(project, project.tasks[0], selected_date="2026-06-02")
    daily_dialog = DailyDialog(project, selected_date="2026-06-02")
    archive_dialog = ArchiveDialog(project)
    inbox_dialog = InboxTaskDialog()

    for widget in [
        project_dialog.deadline,
        task_dialog.start,
        task_dialog.completed,
        daily_dialog.date,
        archive_dialog.date,
        inbox_dialog.created,
    ]:
        assert isinstance(widget, QDateEdit)
        assert widget.calendarPopup()
        calendar = widget.calendarWidget()
        assert isinstance(calendar, QCalendarWidget)
        assert calendar.firstDayOfWeek() == Qt.Monday
        assert calendar.verticalHeaderFormat() == QCalendarWidget.NoVerticalHeader
        assert "qt_calendar_navigationbar" in calendar.styleSheet()

    assert project_dialog.values()["deadline"] == "2026-06-30"
    assert task_dialog.values()["startDate"] == "2026-06-01"
    assert daily_dialog.values()["date"] == "2026-06-02"


def test_daily_dialog_defaults_to_selected_task_and_owner():
    app()
    project = Project(
        name="Project",
        tasks=[
            Task(id="t1", title="Task A", responsible="Owner A", startDate="2026-06-01"),
            Task(id="t2", title="Task B", responsible="Owner B", startDate="2026-06-02"),
        ],
    )

    dialog = DailyDialog(project, selected_date="2026-06-03", selected_task_id="t2")

    assert dialog.values()["taskId"] == "t2"
    assert dialog.values()["responsible"] == "Owner B"


def test_core_business_dialogs_use_modern_base():
    app()
    project = Project(
        name="Project",
        deadline="2026-06-30",
        tasks=[Task(id="t1", title="Task", startDate="2026-06-01")],
    )

    dialogs = [
        ProjectDialog(project),
        TaskDialog(project, project.tasks[0], selected_date="2026-06-02"),
        DailyDialog(project, selected_date="2026-06-02"),
        ArchiveDialog(project),
        InboxTaskDialog(),
    ]

    assert all(isinstance(dialog, AppDialog) for dialog in dialogs)
    assert all(not dialog.windowIcon().isNull() for dialog in dialogs)


def test_main_window_uses_application_icon(monkeypatch):
    window = window_for_workspace(monkeypatch, make_cross_project_workspace())

    assert not window.windowIcon().isNull()


def test_task_detail_layout_has_stable_minimum_heights(monkeypatch):
    window = window_for_workspace(monkeypatch, make_cross_project_workspace())

    assert window.context_tabs.minimumHeight() >= 198
    assert window.detail_title.minimumHeight() >= 24
    assert window.detail_meta.minimumHeight() >= 22
    assert window.detail_progress.height() >= 12
    assert window.detail_note.minimumHeight() >= 46


def test_packaging_scripts_embed_application_icon():
    root = Path(__file__).resolve().parents[1]
    build_exe = (root / "build_exe.bat").read_text(encoding="utf-8")
    build_setup = (root / "build_setup.ps1").read_text(encoding="utf-8")

    assert "--icon" in build_exe
    assert "assets\\project_manage.ico" in build_exe
    assert "--icon $IconPath" in build_setup
    assert "$shortcut.IconLocation" in build_setup


def test_sidebar_navigation_uses_embedded_pages():
    app()
    window = MainWindow()

    for name in ["总览", "项目看板", "待归档任务", "项目档案", "任务表格", "日报记录", "风险看板", "数据中心"]:
        window.navigate_to(name)
        assert window.active_page_name == name
        assert isinstance(window.page_stack, QStackedWidget)
        assert window.page_stack.currentWidget() is window.page_widgets[name]

    window.navigate_to("任务计划")
    assert window.page_stack.currentWidget() is window.page_widgets["任务计划"]


def test_task_ledger_defaults_to_all_projects(monkeypatch):
    window = window_for_workspace(monkeypatch, make_cross_project_workspace())

    window.navigate_to("任务表格")
    table = table_with_headers(window.page_stack.currentWidget(), ["项目", "风险", "任务"])

    assert table.rowCount() == 2
    assert {table.item(row, 0).text() for row in range(table.rowCount())} == {"Alpha project", "Beta project"}


def test_daily_log_page_defaults_to_all_projects(monkeypatch):
    window = window_for_workspace(monkeypatch, make_cross_project_workspace())

    window.navigate_to("日报记录")
    table = table_with_headers(window.page_stack.currentWidget(), ["项目", "日期", "负责人"])

    assert table.rowCount() == 2
    assert {table.item(row, 0).text() for row in range(table.rowCount())} == {"Alpha project", "Beta project"}


def test_inbox_conversion_target_project_is_explicit(monkeypatch):
    window = window_for_workspace(monkeypatch, make_cross_project_workspace())

    window.navigate_to("待归档任务")
    page = window.page_stack.currentWidget()
    combo_texts = [[combo.itemText(index) for index in range(combo.count())] for combo in page.findChildren(QComboBox)]
    button_texts = [button.text() for button in page.findChildren(QPushButton)]

    assert any("Beta project" in texts for texts in combo_texts)
    assert "转为“Beta project”任务" in button_texts
